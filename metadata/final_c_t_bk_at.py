import os
import shutil
from googletrans import Translator
import time
import openpyxl
import datetime

# pip install openpyxl
# !pip install googletrans==4.0.0rc1

# Your existing arrays
folders = ['ta-IN', 'ar', 'bg-BG', 'bn-IN', 'bs-BA', 'ca-ES', 'cs-CZ', 'cy-GB', 'da-DK', 'de-DE', 'el-GR', 'en-GB', 'en-IN', 'en-US', 'es-ES', 'et-EE', 'fi-FI', 'fil-PH', 'fr-FR', 'gu-IN', 'he-IL', 'hi-IN', 'hr-HR', 'hu-HU', 'id-ID', 'is-IS', 'it-IT', 'ja-JP', 'jv-ID', 'km-KH', 'kn-IN', 'ko-KR', 'lv-LV', 'lt-LT', 'ml-IN', 'mr-IN', 'ms-MY', 'nb-NO', 'ne-NP', 'nl-NL', 'pa-IN', 'pl-PL', 'pt-BR', 'pt-PT', 'ro-RO', 'ru-RU', 'si-LK', 'sk-SK', 'sq-AL', 'sr', 'su-ID', 'sv-SE', 'sw-KE', 'te-IN', 'th-TH', 'tr-TR', 'uk-UA', 'ur-PK', 'vi-VN', 'zh-CN']
dest_lang_list = ['ta', 'ar', 'bg', 'bn', 'bs', 'ca', 'cs', 'cy', 'da', 'de', 'el', 'en', 'en', 'en', 'es', 'et', 'fi', 'tl', 'fr', 'gu', 'he', 'hi', 'hr', 'hu', 'id', 'is', 'it', 'ja', 'jw', 'km', 'kn', 'ko', 'lv', 'lt', 'ml', 'mr', 'ms', 'no', 'ne', 'nl', 'pa', 'pl', 'pt', 'pt', 'ro', 'ru', 'si', 'sk', 'sq', 'sr', 'su', 'sv', 'sw', 'te', 'th', 'tr', 'uk', 'ur', 'vi', 'zh-CN']

def create_folders_and_copy(start_index, end_index):
    # Create folders
    selected_folders = folders[start_index:end_index]
    selected_dest_langs = dest_lang_list[start_index:end_index]
    
    for folder in selected_folders:
        os.makedirs(folder, exist_ok=True)
        # time.sleep(4)

    # Get the number of episodes from the user
    num_episodes = int(input("Enter the number of episodes: "))

    # Create episode files in each folder
    for folder in selected_folders:
        for i in range(1, num_episodes + 1):
            open(f"{folder}/{i}.txt", 'w').close()
            # time.sleep(4)

    # Get the input folder path from the user
    input_folder = input("Enter the path of the input folder: ")

    # Copy and paste content to each file
    for folder in selected_folders:
        for i in range(1, num_episodes + 1):
            with open(f"{input_folder}/{i}.txt", 'r') as input_file:
                content = input_file.read()
            with open(f"{folder}/{i}.txt", 'w') as output_file:
                output_file.write(content)
                # time.sleep(4)
    translate_files(selected_folders, selected_dest_langs)            

    print("Process completed.")

def copy_extra_files(start_index, end_index):
    input_folder = input("Enter the path of the input folder: ")
    selected_folders = folders[start_index:end_index]
    
    for folder in selected_folders:
        for file_name in os.listdir(input_folder):
            source_file = os.path.join(input_folder, file_name)
            if os.path.isfile(source_file):
                shutil.copy(source_file, folder)
    translate_files(selected_folders, selected_dest_langs)
    print("Process completed.")

def translate_files(selected_folders, selected_dest_langs):
    translator = Translator()
    home_folder = input("Enter the path to the home folder: ")
    episode_count = int(input("Enter the number of episodes (integer): "))
    folder_sleep_sec = int(input("Enter the number seconds folder need to sleep (integer): "))
    file_sleep_sec = int(input("Enter the number seconds file need to sleep (integer): "))
    for folder_name, dest_lang in zip(selected_folders, selected_dest_langs):
       folder_path = os.path.join(home_folder, folder_name)
   
       if not os.path.exists(folder_path):
           print(f"Folder '{folder_name}' does not exist in '{home_folder}'. Skipping...")
           continue
   
       for episode in range(1, episode_count + 1):
           file_name = f"{episode}.txt"
           file_path = os.path.join(folder_path, file_name)
   
           if not os.path.exists(file_path):
               print(f"File '{file_name}' does not exist in '{folder_name}'. Skipping...")
               continue
   
           with open(file_path, 'r', encoding='utf-8') as file:
               content = file.read()
               time.sleep(file_sleep_sec)
               translated = translator.translate(content, dest=dest_lang).text
   
           with open(file_path, 'w', encoding='utf-8') as file:
               file.write(translated)
       print(f" '{folder_name}' {folders.index(folder_name)} completed",datetime.datetime.now())           
       time.sleep(folder_sleep_sec)



def excell_operation(trans):
    translator = Translator()
    # home_folder = input("Enter the path to the home folder: ")
    # episode_count = int(input("Enter the number of episodes (integer): "))

    if trans==0:

         book_name = input("Enter the Book name: ")
        #  author_name = input("Enter the Author name: ")
        #  folder_sleep_sec = int(input("Enter the number seconds folder need to sleep (integer): "))
         file_sleep_sec = int(input("Enter the number seconds lang need to sleep (integer): "))
        #  data=[book_name]
         wb = openpyxl.Workbook()
         ws = wb.active
          
     
     
         ws.cell(row=1, column=1).value = book_name
         for i, language in enumerate(dest_lang_list):
              translation = translator.translate(book_name, dest=language).text
              time.sleep(file_sleep_sec)
              ws.cell(row=i+2, column=1).value = translation
         wb.save('translated_texts.xlsx')    




    if trans==1:

         book_name = input("Enter the Book name: ")
         author_name = input("Enter the Author name: ")
         folder_sleep_sec = int(input("Enter the number seconds folder need to sleep (integer): "))
         file_sleep_sec = int(input("Enter the number seconds lang need to sleep (integer): "))
         data=[book_name,author_name]
         wb = openpyxl.Workbook()
         ws = wb.active
          
     
     
         for i, text in enumerate(data):
           ws.cell(row=1, column=i + 1).value = text
     
         # Translate each text into each language and write the translations to the corresponding columns
         # translator = Translator()
         for i, text in enumerate(data):
             for j, language in enumerate(dest_lang_list):
                 translation = translator.translate(text, dest=language,src='en').text
                 print(translation)
                 ws.cell(row=j + 2, column=i + 1).value = translation
                 time.sleep(file_sleep_sec)
             time.sleep(folder_sleep_sec)
         wb.save('translated_texts.xlsx')       



# Ask the user for choice
print("0-create new 1-for add extra files 2-for excell operation")
choice = input("Choose an option (0 or 1 or 2): ")

if choice == "0":
    start_index = int(input("Enter the starting index: "))
    end_index = int(input("Enter the ending index: "))
    create_folders_and_copy(start_index, end_index)
elif choice == "1":
    start_index = int(input("Enter the starting index: "))
    end_index = int(input("Enter the ending index: "))
    copy_extra_files(start_index, end_index)
elif choice == "2":
    print("0-name only 1-name and author")
    trans = int(input("Choose an option (0 or 1 or 2): "))
    excell_operation(trans)    
else:
    print("Invalid choice.")


